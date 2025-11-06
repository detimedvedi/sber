"""
Results Export Module for СберИндекс Anomaly Detection System

This module provides functionality to export anomaly detection results
to various formats including CSV, Excel, and visualizations.
"""

import logging
import os
from datetime import datetime
from typing import Dict, Any, List
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np

logger = logging.getLogger(__name__)


class ResultsExporter:
    """
    Exports anomaly detection results to various formats.
    
    Provides methods to:
    - Export master CSV with all anomalies
    - Export summary Excel with multiple sheets
    - Generate visualizations
    - Generate documentation and reports
    """
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize the ResultsExporter.
        
        Args:
            config: Configuration dictionary containing export settings
        """
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self.config = config
        self.output_dir = config.get('export', {}).get('output_dir', 'output')
        self.timestamp_format = config.get('export', {}).get('timestamp_format', '%Y%m%d_%H%M%S')
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        self.logger.info(f"Output directory: {self.output_dir}")
    
    def _get_timestamp(self) -> str:
        """
        Generate timestamp string for filenames.
        
        Returns:
            Formatted timestamp string
        """
        return datetime.now().strftime(self.timestamp_format)
    
    def export_master_csv(self, df: pd.DataFrame, filename: str = None) -> str:
        """
        Export master anomalies table to CSV format.
        
        Creates a comprehensive CSV file containing all detected anomalies
        with proper encoding for Russian text. The CSV includes all required
        columns as specified in the requirements.
        
        Args:
            df: DataFrame containing all anomalies with columns:
                - anomaly_id: Unique identifier for the anomaly
                - territory_id: Municipal territory identifier
                - municipal_name: Name of municipality
                - region_name: Name of region
                - indicator: Name of the indicator showing anomaly
                - anomaly_type: Type of anomaly detected
                - actual_value: Actual value of the indicator
                - expected_value: Expected value (may be None)
                - deviation: Absolute deviation from expected
                - deviation_pct: Percentage deviation
                - severity_score: Severity score (0-100)
                - z_score: Z-score (may be None)
                - data_source: Source of data ('sberindex' or 'rosstat')
                - detection_method: Method used to detect anomaly
                - description: Human-readable description
                - detected_at: Timestamp of detection
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created CSV file
        """
        self.logger.info("Exporting master anomalies CSV")
        
        if df.empty:
            self.logger.warning("No anomalies to export")
            # Create empty CSV with headers
            df = pd.DataFrame(columns=[
                'anomaly_id', 'territory_id', 'municipal_name', 'region_name',
                'indicator', 'anomaly_type', 'actual_value', 'expected_value',
                'deviation', 'deviation_pct', 'severity_score', 'z_score',
                'data_source', 'detection_method', 'description', 'detected_at'
            ])
        
        # Generate filename with timestamp
        if filename is None:
            timestamp = self._get_timestamp()
            filename = f"anomalies_master_{timestamp}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.csv")
        
        # Select and order columns for export
        export_columns = [
            'anomaly_id',
            'territory_id',
            'municipal_name',
            'region_name',
            'indicator',
            'anomaly_type',
            'actual_value',
            'expected_value',
            'deviation',
            'deviation_pct',
            'severity_score',
            'z_score',
            'data_source',
            'detection_method',
            'description',
            'detected_at'
        ]
        
        # Filter to only include columns that exist in the DataFrame
        available_columns = [col for col in export_columns if col in df.columns]
        
        if len(available_columns) < len(export_columns):
            missing_columns = set(export_columns) - set(available_columns)
            self.logger.warning(f"Missing columns in export: {missing_columns}")
        
        export_df = df[available_columns].copy()
        
        # Format datetime columns if present
        if 'detected_at' in export_df.columns:
            export_df['detected_at'] = pd.to_datetime(export_df['detected_at']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Round numeric columns for readability
        numeric_columns = ['actual_value', 'expected_value', 'deviation', 'deviation_pct', 
                          'severity_score', 'z_score']
        for col in numeric_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(2)
        
        # Export to CSV with UTF-8 encoding for Russian text
        export_df.to_csv(
            filepath,
            index=False,
            encoding='utf-8-sig',  # UTF-8 with BOM for Excel compatibility
            float_format='%.2f'
        )
        
        self.logger.info(f"Exported {len(export_df)} anomalies to {filepath}")
        self.logger.info(f"File size: {os.path.getsize(filepath) / 1024:.2f} KB")
        
        return filepath
    
    def export_summary_excel(
        self,
        all_anomalies: pd.DataFrame,
        categorized_anomalies: Dict[str, pd.DataFrame],
        municipality_scores: pd.DataFrame,
        summary_stats: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Export comprehensive Excel file with multiple sheets.
        
        Creates an Excel workbook with multiple sheets containing:
        - Executive_Summary: Management-friendly summary with key findings and top 10 municipalities
        - Overview: Summary statistics and key findings
        - Statistical_Outliers: Statistical outlier anomalies
        - Temporal_Anomalies: Temporal anomaly detections
        - Geographic_Anomalies: Geographic anomaly detections
        - Cross_Source_Discrepancies: Cross-source comparison issues
        - Logical_Inconsistencies: Logical consistency problems
        - Data_Quality_Issues: Data quality problems
        - Top_Anomalous_Municipalities: Top 50 municipalities by severity
        - Data_Dictionary: Explanation of all columns and anomaly types
        
        Args:
            all_anomalies: DataFrame with all detected anomalies
            categorized_anomalies: Dictionary of anomalies grouped by type
            municipality_scores: DataFrame with municipality-level scores
            summary_stats: Dictionary with summary statistics
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created Excel file
        """
        self.logger.info("Exporting summary Excel with multiple sheets")
        
        # Generate filename with timestamp
        if filename is None:
            timestamp = self._get_timestamp()
            filename = f"anomalies_summary_{timestamp}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            
            # Sheet 1: Executive Summary (NEW)
            self._create_executive_summary_sheet(writer, all_anomalies, municipality_scores)
            
            # Sheet 2: Overview
            self._create_overview_sheet(writer, summary_stats, all_anomalies, municipality_scores)
            
            # Sheet 3-8: Anomaly type sheets
            self._create_anomaly_type_sheets(writer, categorized_anomalies)
            
            # Sheet 9: Top Anomalous Municipalities
            self._create_top_municipalities_sheet(writer, municipality_scores)
            
            # Sheet 10: Data Dictionary
            self._create_data_dictionary_sheet(writer)
        
        self.logger.info(f"Exported Excel summary to {filepath}")
        self.logger.info(f"File size: {os.path.getsize(filepath) / 1024:.2f} KB")
        
        return filepath
    
    def _create_overview_sheet(
        self,
        writer: pd.ExcelWriter,
        summary_stats: Dict[str, Any],
        all_anomalies: pd.DataFrame,
        municipality_scores: pd.DataFrame
    ):
        """
        Create Overview sheet with summary statistics.
        
        Args:
            writer: Excel writer object
            summary_stats: Dictionary with summary statistics
            all_anomalies: DataFrame with all anomalies
            municipality_scores: DataFrame with municipality scores
        """
        self.logger.info("Creating Overview sheet")
        
        # Create overview data
        overview_data = []
        
        # General statistics
        overview_data.append(['GENERAL STATISTICS', ''])
        overview_data.append(['Total Anomalies Detected', summary_stats.get('total_anomalies', 0)])
        overview_data.append(['Total Municipalities Affected', summary_stats.get('total_municipalities_affected', 0)])
        overview_data.append(['', ''])
        
        # Severity statistics
        overview_data.append(['SEVERITY STATISTICS', ''])
        severity_stats = summary_stats.get('severity_stats', {})
        overview_data.append(['Mean Severity Score', f"{severity_stats.get('mean', 0):.2f}"])
        overview_data.append(['Median Severity Score', f"{severity_stats.get('median', 0):.2f}"])
        overview_data.append(['Min Severity Score', f"{severity_stats.get('min', 0):.2f}"])
        overview_data.append(['Max Severity Score', f"{severity_stats.get('max', 0):.2f}"])
        overview_data.append(['Std Dev Severity Score', f"{severity_stats.get('std', 0):.2f}"])
        overview_data.append(['', ''])
        
        # Severity distribution
        overview_data.append(['SEVERITY DISTRIBUTION', ''])
        severity_dist = summary_stats.get('severity_distribution', {})
        for category, count in severity_dist.items():
            overview_data.append([str(category), count])
        overview_data.append(['', ''])
        
        # Anomalies by type
        overview_data.append(['ANOMALIES BY TYPE', ''])
        by_type = summary_stats.get('by_type', {})
        for anomaly_type, count in sorted(by_type.items(), key=lambda x: x[1], reverse=True):
            overview_data.append([anomaly_type, count])
        overview_data.append(['', ''])
        
        # Data source distribution
        overview_data.append(['DATA SOURCE DISTRIBUTION', ''])
        data_source_dist = summary_stats.get('data_source_distribution', {})
        for source, count in data_source_dist.items():
            overview_data.append([source, count])
        overview_data.append(['', ''])
        
        # Top 10 regions by anomaly count
        overview_data.append(['TOP 10 REGIONS BY ANOMALY COUNT', ''])
        by_region = summary_stats.get('by_region', {})
        for region, count in list(by_region.items())[:10]:
            overview_data.append([region, count])
        overview_data.append(['', ''])
        
        # Top 10 municipalities
        overview_data.append(['TOP 10 MOST ANOMALOUS MUNICIPALITIES', ''])
        overview_data.append(['Rank', 'Municipality', 'Region', 'Total Anomalies', 'Total Severity Score'])
        if not municipality_scores.empty:
            top_10 = municipality_scores.head(10)
            for _, row in top_10.iterrows():
                overview_data.append([
                    row.get('rank', ''),
                    row.get('municipal_name', ''),
                    row.get('region_name', ''),
                    row.get('total_anomalies_count', 0),
                    f"{row.get('total_severity_score', 0):.2f}"
                ])
        
        # Create DataFrame and write to Excel
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='Overview', index=False, header=False)
        
        self.logger.info("Overview sheet created")
    
    def _create_executive_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        all_anomalies: pd.DataFrame,
        municipality_scores: pd.DataFrame
    ):
        """
        Create Executive Summary sheet for management reporting.
        
        Creates a management-friendly summary sheet with:
        - Key metrics and statistics
        - Top 10 municipalities table
        - Key findings as bullet points
        - Critical anomalies highlighted in red
        
        Args:
            writer: Excel writer object
            all_anomalies: DataFrame with all anomalies
            municipality_scores: DataFrame with municipality scores
        """
        self.logger.info("Creating Executive Summary sheet")
        
        # Generate executive summary using ExecutiveSummaryGenerator
        summary_generator = ExecutiveSummaryGenerator(self.config)
        executive_summary = summary_generator.generate(all_anomalies)
        
        # Create summary data for Excel
        summary_data = []
        
        # Title
        summary_data.append(['EXECUTIVE SUMMARY / РЕЗЮМЕ ДЛЯ РУКОВОДСТВА', ''])
        summary_data.append(['', ''])
        
        # Key Metrics
        summary_data.append(['КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ', ''])
        summary_data.append(['Всего обнаружено аномалий', executive_summary['total_anomalies']])
        summary_data.append(['Критических аномалий (severity > 90)', executive_summary['critical_count']])
        summary_data.append(['Затронуто муниципальных образований', executive_summary['affected_municipalities']])
        summary_data.append(['', ''])
        
        # Key Findings
        summary_data.append(['ОСНОВНЫЕ ВЫВОДЫ', ''])
        for i, finding in enumerate(executive_summary['key_findings'], 1):
            summary_data.append([f'{i}.', finding])
        summary_data.append(['', ''])
        
        # Recommendations
        summary_data.append(['РЕКОМЕНДАЦИИ', ''])
        for i, recommendation in enumerate(executive_summary['recommendations'], 1):
            summary_data.append([f'{i}.', recommendation])
        summary_data.append(['', ''])
        summary_data.append(['', ''])
        
        # Top 10 Municipalities Table
        summary_data.append(['ТОП-10 МУНИЦИПАЛИТЕТОВ ПО УРОВНЮ РИСКА', ''])
        summary_data.append(['', ''])
        
        # Table headers
        summary_data.append([
            'Ранг',
            'Муниципалитет',
            'Регион',
            'Количество аномалий',
            'Критических',
            'Общая серьезность',
            'Макс. серьезность',
            'Типы аномалий'
        ])
        
        # Table data
        top_10 = executive_summary['top_10_municipalities']
        for i, municipality in enumerate(top_10, 1):
            # Format anomaly types
            anomaly_types_str = ', '.join(municipality.get('anomaly_types', []))[:50]  # Limit length
            
            summary_data.append([
                i,
                municipality.get('municipal_name', ''),
                municipality.get('region_name', ''),
                municipality.get('anomaly_count', 0),
                municipality.get('critical_count', 0),
                f"{municipality.get('total_severity', 0):.1f}",
                f"{municipality.get('max_severity', 0):.1f}",
                anomaly_types_str
            ])
        
        # Create DataFrame and write to Excel
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False, header=False)
        
        # Get the workbook and worksheet to apply formatting
        workbook = writer.book
        worksheet = writer.sheets['Executive_Summary']
        
        # Apply formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title formatting (row 1)
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        worksheet['A1'].font = title_font
        worksheet['A1'].fill = title_fill
        worksheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
        worksheet.merge_cells('A1:B1')
        
        # Section headers formatting
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Find section header rows
        section_headers = [
            ('КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ', 3),
            ('ОСНОВНЫЕ ВЫВОДЫ', None),
            ('РЕКОМЕНДАЦИИ', None),
            ('ТОП-10 МУНИЦИПАЛИТЕТОВ ПО УРОВНЮ РИСКА', None)
        ]
        
        current_row = 1
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            current_row += 1
            cell_value = row[0].value
            if cell_value in [header[0] for header in section_headers]:
                row[0].font = header_font
                row[0].fill = header_fill
                row[0].alignment = Alignment(horizontal='left', vertical='center')
                worksheet.merge_cells(f'A{current_row}:B{current_row}')
        
        # Format top 10 table header
        table_start_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), 1):
            if row[0].value == 'Ранг':
                table_start_row = row_idx
                break
        
        if table_start_row:
            # Header row formatting
            table_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            table_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_idx in range(1, 9):  # 8 columns
                cell = worksheet.cell(row=table_start_row, column=col_idx)
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows formatting
            for row_idx in range(table_start_row + 1, table_start_row + 11):  # Top 10 rows
                if row_idx > worksheet.max_row:
                    break
                
                for col_idx in range(1, 9):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Highlight critical anomalies in red
                    critical_count_col = 5  # Column E (Критических)
                    if col_idx == critical_count_col:
                        critical_count = cell.value if isinstance(cell.value, (int, float)) else 0
                        if critical_count > 0:
                            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')
                    
                    # Highlight high severity in red
                    max_severity_col = 7  # Column G (Макс. серьезность)
                    if col_idx == max_severity_col:
                        try:
                            severity_value = float(str(cell.value).replace(',', '.')) if cell.value else 0
                            if severity_value > 90:
                                cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')
                                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        except (ValueError, AttributeError):
                            pass
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 60
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 18
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 40
        
        # Set row heights for better readability
        for row_idx in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
        
        # Make title row taller
        worksheet.row_dimensions[1].height = 30
        
        self.logger.info("Executive Summary sheet created with formatting")
    
    def _create_anomaly_type_sheets(
        self,
        writer: pd.ExcelWriter,
        categorized_anomalies: Dict[str, pd.DataFrame]
    ):
        """
        Create sheets for each anomaly type with descriptive statistics.
        
        Args:
            writer: Excel writer object
            categorized_anomalies: Dictionary of anomalies grouped by type
        """
        self.logger.info("Creating anomaly type sheets")
        
        # Define sheet names mapping
        sheet_names = {
            'statistical_outliers': 'Statistical_Outliers',
            'temporal_anomalies': 'Temporal_Anomalies',
            'geographic_anomalies': 'Geographic_Anomalies',
            'cross_source_discrepancies': 'Cross_Source_Discrepancies',
            'logical_inconsistencies': 'Logical_Inconsistencies',
            'data_quality_issues': 'Data_Quality_Issues'
        }
        
        # Export columns for anomaly sheets
        export_columns = [
            'territory_id',
            'municipal_name',
            'region_name',
            'indicator',
            'actual_value',
            'expected_value',
            'deviation',
            'deviation_pct',
            'severity_score',
            'z_score',
            'data_source',
            'detection_method',
            'description'
        ]
        
        for category, sheet_name in sheet_names.items():
            df = categorized_anomalies.get(category, pd.DataFrame())
            
            if df.empty:
                # Create empty sheet with headers
                empty_df = pd.DataFrame(columns=export_columns)
                empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.info(f"Created empty sheet: {sheet_name}")
            else:
                # Filter to available columns
                available_columns = [col for col in export_columns if col in df.columns]
                export_df = df[available_columns].copy()
                
                # Round numeric columns
                numeric_columns = ['actual_value', 'expected_value', 'deviation', 
                                 'deviation_pct', 'severity_score', 'z_score']
                for col in numeric_columns:
                    if col in export_df.columns:
                        export_df[col] = export_df[col].round(2)
                
                # Write data to sheet
                export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Add descriptive statistics at the bottom
                start_row = len(export_df) + 3
                
                # Create statistics summary
                stats_data = []
                stats_data.append(['DESCRIPTIVE STATISTICS', ''])
                stats_data.append(['Total Count', len(export_df)])
                stats_data.append(['Unique Municipalities', export_df['territory_id'].nunique() if 'territory_id' in export_df.columns else 0])
                
                if 'severity_score' in export_df.columns:
                    stats_data.append(['Mean Severity', f"{export_df['severity_score'].mean():.2f}"])
                    stats_data.append(['Median Severity', f"{export_df['severity_score'].median():.2f}"])
                    stats_data.append(['Max Severity', f"{export_df['severity_score'].max():.2f}"])
                    stats_data.append(['Min Severity', f"{export_df['severity_score'].min():.2f}"])
                
                if 'data_source' in export_df.columns:
                    stats_data.append(['', ''])
                    stats_data.append(['BY DATA SOURCE', ''])
                    source_counts = export_df['data_source'].value_counts()
                    for source, count in source_counts.items():
                        stats_data.append([source, count])
                
                # Write statistics to sheet
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=False,
                    header=False
                )
                
                self.logger.info(f"Created sheet: {sheet_name} with {len(export_df)} anomalies")
    
    def _create_top_municipalities_sheet(
        self,
        writer: pd.ExcelWriter,
        municipality_scores: pd.DataFrame
    ):
        """
        Create Top Anomalous Municipalities sheet.
        
        Args:
            writer: Excel writer object
            municipality_scores: DataFrame with municipality scores
        """
        self.logger.info("Creating Top Anomalous Municipalities sheet")
        
        # Get top N municipalities from config
        top_n = self.config.get('export', {}).get('top_n_municipalities', 50)
        
        if municipality_scores.empty:
            # Create empty sheet
            empty_df = pd.DataFrame(columns=[
                'rank', 'territory_id', 'municipal_name', 'region_name',
                'total_anomalies_count', 'total_severity_score',
                'average_severity_score', 'max_severity', 'anomaly_types'
            ])
            empty_df.to_excel(writer, sheet_name='Top_Anomalous_Municipalities', index=False)
            self.logger.info("Created empty Top Anomalous Municipalities sheet")
        else:
            # Get top N municipalities
            top_municipalities = municipality_scores.head(top_n).copy()
            
            # Format anomaly_types list as string
            if 'anomaly_types' in top_municipalities.columns:
                top_municipalities['anomaly_types'] = top_municipalities['anomaly_types'].apply(
                    lambda x: ', '.join(x) if isinstance(x, list) else str(x)
                )
            
            # Round numeric columns
            numeric_columns = ['total_severity_score', 'average_severity_score', 'max_severity']
            for col in numeric_columns:
                if col in top_municipalities.columns:
                    top_municipalities[col] = top_municipalities[col].round(2)
            
            # Write to Excel
            top_municipalities.to_excel(
                writer,
                sheet_name='Top_Anomalous_Municipalities',
                index=False
            )
            
            self.logger.info(f"Created Top Anomalous Municipalities sheet with {len(top_municipalities)} entries")
    
    def _create_data_dictionary_sheet(self, writer: pd.ExcelWriter):
        """
        Create Data Dictionary sheet explaining all columns and anomaly types.
        
        Args:
            writer: Excel writer object
        """
        self.logger.info("Creating Data Dictionary sheet")
        
        # Column definitions
        column_definitions = [
            ['COLUMN DEFINITIONS', ''],
            ['', ''],
            ['Column Name', 'Description'],
            ['anomaly_id', 'Unique identifier for the anomaly (UUID)'],
            ['territory_id', 'Municipal territory identifier'],
            ['municipal_name', 'Name of the municipality'],
            ['region_name', 'Name of the region/oblast'],
            ['indicator', 'Name of the indicator showing anomaly'],
            ['anomaly_type', 'Type of anomaly detected (see Anomaly Types section)'],
            ['actual_value', 'Actual value of the indicator'],
            ['expected_value', 'Expected value based on statistical analysis (may be null)'],
            ['deviation', 'Absolute deviation from expected value'],
            ['deviation_pct', 'Percentage deviation from expected value'],
            ['severity_score', 'Severity score from 0 to 100 (higher = more severe)'],
            ['z_score', 'Statistical z-score (number of standard deviations from mean)'],
            ['data_source', 'Source of data: "sberindex" or "rosstat"'],
            ['detection_method', 'Method used to detect the anomaly'],
            ['description', 'Human-readable description of the anomaly'],
            ['detected_at', 'Timestamp when anomaly was detected'],
            ['rank', 'Overall rank by severity score'],
            ['total_anomalies_count', 'Total number of anomalies for a municipality'],
            ['total_severity_score', 'Sum of all severity scores for a municipality'],
            ['average_severity_score', 'Average severity score for a municipality'],
            ['max_severity', 'Maximum severity score among all anomalies for a municipality'],
            ['anomaly_types', 'List of anomaly types present for a municipality'],
            ['', ''],
            ['', ''],
            ['ANOMALY TYPES', ''],
            ['', ''],
            ['Type', 'Description'],
            ['statistical_outlier', 'Value significantly deviates from statistical norms (z-score, IQR, percentiles)'],
            ['temporal_anomaly', 'Unusual temporal patterns: sudden spikes, drops, trend reversals, high volatility'],
            ['geographic_anomaly', 'Value differs significantly from regional average or neighboring municipalities'],
            ['cross_source_discrepancy', 'Large discrepancy between СберИндекс and Росстат data for same indicator'],
            ['logical_inconsistency', 'Logically impossible or contradictory values (e.g., negative where positive expected)'],
            ['data_quality_issue', 'Missing data, duplicates, or other data quality problems'],
            ['', ''],
            ['', ''],
            ['SEVERITY SCORE INTERPRETATION', ''],
            ['', ''],
            ['Score Range', 'Interpretation'],
            ['0-25', 'Low severity - Minor deviation, may be normal variation'],
            ['25-50', 'Medium severity - Notable deviation, worth investigating'],
            ['50-75', 'High severity - Significant deviation, likely requires attention'],
            ['75-100', 'Critical severity - Extreme deviation, requires immediate investigation'],
            ['', ''],
            ['', ''],
            ['DETECTION METHODS', ''],
            ['', ''],
            ['Method', 'Description'],
            ['z_score', 'Identifies values more than 3 standard deviations from mean'],
            ['iqr', 'Identifies values outside 1.5 * IQR from quartiles'],
            ['percentile', 'Identifies values in top/bottom 1% of distribution'],
            ['sudden_spike', 'Detects period-over-period growth >100% or drops <-50%'],
            ['trend_reversal', 'Detects significant changes in trend direction'],
            ['high_volatility', 'Detects volatility >2x median volatility'],
            ['regional_outlier', 'Detects values >2 std dev from regional mean'],
            ['cluster_outlier', 'Detects municipalities differing from similar neighbors'],
            ['correlation_check', 'Checks correlation between СберИндекс and Росстат'],
            ['discrepancy_check', 'Identifies >50% differences between data sources'],
            ['negative_value_check', 'Identifies negative values where only positive expected'],
            ['impossible_ratio_check', 'Identifies logically impossible ratios'],
            ['contradiction_check', 'Identifies contradictory indicators'],
            ['', ''],
            ['', ''],
            ['DATA SOURCES', ''],
            ['', ''],
            ['Source', 'Description'],
            ['sberindex', 'СберИндекс data: connection, consumption, market_access indicators'],
            ['rosstat', 'Росстат official statistics: population, migration, salary data'],
            ['', ''],
            ['', ''],
            ['NOTES', ''],
            ['', ''],
            ['Note', 'Details'],
            ['Multiple anomalies', 'A single municipality may have multiple anomalies across different indicators'],
            ['Severity calculation', 'Severity scores are calculated based on deviation magnitude and statistical significance'],
            ['Expected values', 'Expected values may be null for some anomaly types (e.g., logical inconsistencies)'],
            ['Interpretation', 'Anomalies indicate potential data quality issues or genuinely unusual patterns requiring investigation'],
            ['False positives', 'Some anomalies may be legitimate unusual cases rather than errors'],
        ]
        
        # Create DataFrame and write to Excel
        dict_df = pd.DataFrame(column_definitions)
        dict_df.to_excel(writer, sheet_name='Data_Dictionary', index=False, header=False)
        
        self.logger.info("Data Dictionary sheet created")
    
    def generate_visualizations(
        self,
        all_anomalies: pd.DataFrame,
        municipality_scores: pd.DataFrame,
        summary_stats: Dict[str, Any],
        filename_prefix: str = None
    ) -> Dict[str, str]:
        """
        Generate visualizations for anomaly detection results.
        
        Creates the following visualizations:
        1. Bar chart for anomaly distribution by type
        2. Horizontal bar chart for top 20 municipalities by severity
        3. Heatmap for geographic distribution by region
        4. Histogram for severity score distribution
        
        All visualizations are saved as PNG files in the output directory.
        
        Args:
            all_anomalies: DataFrame with all detected anomalies
            municipality_scores: DataFrame with municipality-level scores
            summary_stats: Dictionary with summary statistics
            filename_prefix: Optional prefix for filenames
            
        Returns:
            Dictionary mapping visualization type to filepath
        """
        self.logger.info("Generating visualizations")
        
        # Get visualization config
        viz_config = self.config.get('visualization', {})
        figure_size = tuple(viz_config.get('figure_size', [12, 8]))
        dpi = viz_config.get('dpi', 300)
        style = viz_config.get('style', 'seaborn-v0_8')
        
        # Set matplotlib style
        try:
            plt.style.use(style)
        except:
            self.logger.warning(f"Style '{style}' not available, using default")
            plt.style.use('default')
        
        # Generate timestamp for filenames
        if filename_prefix is None:
            timestamp = self._get_timestamp()
            filename_prefix = f"viz_{timestamp}"
        
        # Dictionary to store generated file paths
        generated_files = {}
        
        # 1. Bar chart for anomaly distribution by type
        try:
            filepath = self._create_anomaly_type_distribution_chart(
                all_anomalies,
                summary_stats,
                filename_prefix,
                figure_size,
                dpi
            )
            generated_files['anomaly_type_distribution'] = filepath
        except Exception as e:
            self.logger.error(f"Failed to create anomaly type distribution chart: {e}")
        
        # 2. Horizontal bar chart for top 20 municipalities
        try:
            filepath = self._create_top_municipalities_chart(
                municipality_scores,
                filename_prefix,
                figure_size,
                dpi
            )
            generated_files['top_municipalities'] = filepath
        except Exception as e:
            self.logger.error(f"Failed to create top municipalities chart: {e}")
        
        # 3. Heatmap for geographic distribution by region
        try:
            filepath = self._create_geographic_heatmap(
                all_anomalies,
                summary_stats,
                filename_prefix,
                figure_size,
                dpi
            )
            generated_files['geographic_heatmap'] = filepath
        except Exception as e:
            self.logger.error(f"Failed to create geographic heatmap: {e}")
        
        # 4. Histogram for severity score distribution
        try:
            filepath = self._create_severity_distribution_histogram(
                all_anomalies,
                filename_prefix,
                figure_size,
                dpi
            )
            generated_files['severity_distribution'] = filepath
        except Exception as e:
            self.logger.error(f"Failed to create severity distribution histogram: {e}")
        
        self.logger.info(f"Generated {len(generated_files)} visualizations")
        return generated_files
    
    def _create_anomaly_type_distribution_chart(
        self,
        all_anomalies: pd.DataFrame,
        summary_stats: Dict[str, Any],
        filename_prefix: str,
        figure_size: tuple,
        dpi: int
    ) -> str:
        """
        Create bar chart showing distribution of anomalies by type.
        
        Args:
            all_anomalies: DataFrame with all anomalies
            summary_stats: Dictionary with summary statistics
            filename_prefix: Prefix for filename
            figure_size: Figure size tuple (width, height)
            dpi: DPI for output image
            
        Returns:
            Path to saved PNG file
        """
        self.logger.info("Creating anomaly type distribution chart")
        
        # Get anomaly counts by type
        by_type = summary_stats.get('by_type', {})
        
        if not by_type:
            self.logger.warning("No anomaly type data available")
            # Create empty chart
            by_type = {'No Data': 0}
        
        # Sort by count descending
        sorted_types = sorted(by_type.items(), key=lambda x: x[1], reverse=True)
        types = [t[0] for t in sorted_types]
        counts = [t[1] for t in sorted_types]
        
        # Create figure
        fig, ax = plt.subplots(figsize=figure_size)
        
        # Create bar chart
        bars = ax.bar(range(len(types)), counts, color='steelblue', alpha=0.8, edgecolor='black')
        
        # Customize chart
        ax.set_xlabel('Anomaly Type', fontsize=12, fontweight='bold')
        ax.set_ylabel('Number of Anomalies', fontsize=12, fontweight='bold')
        ax.set_title('Distribution of Anomalies by Type', fontsize=14, fontweight='bold', pad=20)
        ax.set_xticks(range(len(types)))
        ax.set_xticklabels(types, rotation=45, ha='right')
        
        # Add value labels on bars
        for i, (bar, count) in enumerate(zip(bars, counts)):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(count)}',
                   ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        # Add grid for readability
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save figure
        filename = f"{filename_prefix}_anomaly_type_distribution.png"
        filepath = os.path.join(self.output_dir, filename)
        plt.savefig(filepath, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        self.logger.info(f"Saved anomaly type distribution chart to {filepath}")
        return filepath
    
    def _create_top_municipalities_chart(
        self,
        municipality_scores: pd.DataFrame,
        filename_prefix: str,
        figure_size: tuple,
        dpi: int
    ) -> str:
        """
        Create horizontal bar chart showing top 20 municipalities by severity.
        
        Args:
            municipality_scores: DataFrame with municipality scores
            filename_prefix: Prefix for filename
            figure_size: Figure size tuple (width, height)
            dpi: DPI for output image
            
        Returns:
            Path to saved PNG file
        """
        self.logger.info("Creating top municipalities chart")
        
        if municipality_scores.empty:
            self.logger.warning("No municipality scores available")
            # Create empty chart
            fig, ax = plt.subplots(figsize=figure_size)
            ax.text(0.5, 0.5, 'No Data Available', ha='center', va='center', fontsize=16)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
        else:
            # Get top 20 municipalities
            top_20 = municipality_scores.head(20).copy()
            
            # Prepare data
            municipalities = top_20['municipal_name'].tolist()
            scores = top_20['total_severity_score'].tolist()
            
            # Reverse order for horizontal bar chart (highest at top)
            municipalities = municipalities[::-1]
            scores = scores[::-1]
            
            # Create figure
            fig, ax = plt.subplots(figsize=figure_size)
            
            # Create horizontal bar chart
            bars = ax.barh(range(len(municipalities)), scores, color='coral', alpha=0.8, edgecolor='black')
            
            # Customize chart
            ax.set_xlabel('Total Severity Score', fontsize=12, fontweight='bold')
            ax.set_ylabel('Municipality', fontsize=12, fontweight='bold')
            ax.set_title('Top 20 Most Anomalous Municipalities', fontsize=14, fontweight='bold', pad=20)
            ax.set_yticks(range(len(municipalities)))
            ax.set_yticklabels(municipalities, fontsize=9)
            
            # Add value labels on bars
            for i, (bar, score) in enumerate(zip(bars, scores)):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f'{score:.1f}',
                       ha='left', va='center', fontsize=8, fontweight='bold')
            
            # Add grid for readability
            ax.grid(axis='x', alpha=0.3, linestyle='--')
            ax.set_axisbelow(True)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save figure
        filename = f"{filename_prefix}_top_municipalities.png"
        filepath = os.path.join(self.output_dir, filename)
        plt.savefig(filepath, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        self.logger.info(f"Saved top municipalities chart to {filepath}")
        return filepath
    
    def _create_geographic_heatmap(
        self,
        all_anomalies: pd.DataFrame,
        summary_stats: Dict[str, Any],
        filename_prefix: str,
        figure_size: tuple,
        dpi: int
    ) -> str:
        """
        Create heatmap showing geographic distribution of anomalies by region.
        
        Args:
            all_anomalies: DataFrame with all anomalies
            summary_stats: Dictionary with summary statistics
            filename_prefix: Prefix for filename
            figure_size: Figure size tuple (width, height)
            dpi: DPI for output image
            
        Returns:
            Path to saved PNG file
        """
        self.logger.info("Creating geographic heatmap")
        
        # Get anomaly counts by region
        by_region = summary_stats.get('by_region', {})
        
        if not by_region or len(by_region) == 0:
            self.logger.warning("No regional data available")
            # Create empty chart
            fig, ax = plt.subplots(figsize=figure_size)
            ax.text(0.5, 0.5, 'No Regional Data Available', ha='center', va='center', fontsize=16)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
        else:
            # Sort regions by anomaly count
            sorted_regions = sorted(by_region.items(), key=lambda x: x[1], reverse=True)
            
            # Take top 30 regions for better visualization
            top_regions = sorted_regions[:30]
            regions = [r[0] for r in top_regions]
            counts = [r[1] for r in top_regions]
            
            # Create a matrix for heatmap (single column)
            data_matrix = np.array(counts).reshape(-1, 1)
            
            # Create figure with adjusted size for single column heatmap
            fig, ax = plt.subplots(figsize=(8, max(10, len(regions) * 0.4)))
            
            # Create heatmap
            sns.heatmap(
                data_matrix,
                annot=True,
                fmt='g',
                cmap='YlOrRd',
                cbar_kws={'label': 'Number of Anomalies'},
                yticklabels=regions,
                xticklabels=['Anomaly Count'],
                linewidths=0.5,
                linecolor='gray',
                ax=ax
            )
            
            # Customize chart
            ax.set_title('Geographic Distribution of Anomalies by Region (Top 30)', 
                        fontsize=14, fontweight='bold', pad=20)
            ax.set_ylabel('Region', fontsize=12, fontweight='bold')
            ax.set_xlabel('')
            
            # Rotate y-axis labels for better readability
            plt.yticks(rotation=0, fontsize=9)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save figure
        filename = f"{filename_prefix}_geographic_heatmap.png"
        filepath = os.path.join(self.output_dir, filename)
        plt.savefig(filepath, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        self.logger.info(f"Saved geographic heatmap to {filepath}")
        return filepath
    
    def _create_severity_distribution_histogram(
        self,
        all_anomalies: pd.DataFrame,
        filename_prefix: str,
        figure_size: tuple,
        dpi: int
    ) -> str:
        """
        Create histogram showing distribution of severity scores.
        
        Args:
            all_anomalies: DataFrame with all anomalies
            filename_prefix: Prefix for filename
            figure_size: Figure size tuple (width, height)
            dpi: DPI for output image
            
        Returns:
            Path to saved PNG file
        """
        self.logger.info("Creating severity distribution histogram")
        
        if all_anomalies.empty or 'severity_score' not in all_anomalies.columns:
            self.logger.warning("No severity score data available")
            # Create empty chart
            fig, ax = plt.subplots(figsize=figure_size)
            ax.text(0.5, 0.5, 'No Severity Data Available', ha='center', va='center', fontsize=16)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
        else:
            # Get severity scores
            severity_scores = all_anomalies['severity_score'].dropna()
            
            if len(severity_scores) == 0:
                self.logger.warning("No valid severity scores")
                fig, ax = plt.subplots(figsize=figure_size)
                ax.text(0.5, 0.5, 'No Valid Severity Scores', ha='center', va='center', fontsize=16)
                ax.set_xlim(0, 1)
                ax.set_ylim(0, 1)
                ax.axis('off')
            else:
                # Create figure
                fig, ax = plt.subplots(figsize=figure_size)
                
                # Create histogram
                n, bins, patches = ax.hist(
                    severity_scores,
                    bins=20,
                    color='mediumseagreen',
                    alpha=0.7,
                    edgecolor='black',
                    linewidth=1.2
                )
                
                # Color bars by severity level
                # Low (0-25): green, Medium (25-50): yellow, High (50-75): orange, Critical (75-100): red
                for i, patch in enumerate(patches):
                    bin_center = (bins[i] + bins[i+1]) / 2
                    if bin_center < 25:
                        patch.set_facecolor('lightgreen')
                    elif bin_center < 50:
                        patch.set_facecolor('gold')
                    elif bin_center < 75:
                        patch.set_facecolor('orange')
                    else:
                        patch.set_facecolor('tomato')
                
                # Customize chart
                ax.set_xlabel('Severity Score', fontsize=12, fontweight='bold')
                ax.set_ylabel('Frequency', fontsize=12, fontweight='bold')
                ax.set_title('Distribution of Anomaly Severity Scores', fontsize=14, fontweight='bold', pad=20)
                
                # Add vertical lines for severity thresholds
                ax.axvline(x=25, color='green', linestyle='--', linewidth=1.5, alpha=0.7, label='Low/Medium')
                ax.axvline(x=50, color='orange', linestyle='--', linewidth=1.5, alpha=0.7, label='Medium/High')
                ax.axvline(x=75, color='red', linestyle='--', linewidth=1.5, alpha=0.7, label='High/Critical')
                
                # Add statistics text box
                stats_text = f'Mean: {severity_scores.mean():.2f}\n'
                stats_text += f'Median: {severity_scores.median():.2f}\n'
                stats_text += f'Std Dev: {severity_scores.std():.2f}\n'
                stats_text += f'Count: {len(severity_scores)}'
                
                ax.text(0.98, 0.97, stats_text,
                       transform=ax.transAxes,
                       fontsize=10,
                       verticalalignment='top',
                       horizontalalignment='right',
                       bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
                
                # Add legend
                ax.legend(loc='upper left', fontsize=9)
                
                # Add grid for readability
                ax.grid(axis='y', alpha=0.3, linestyle='--')
                ax.set_axisbelow(True)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save figure
        filename = f"{filename_prefix}_severity_distribution.png"
        filepath = os.path.join(self.output_dir, filename)
        plt.savefig(filepath, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        self.logger.info(f"Saved severity distribution histogram to {filepath}")
        return filepath

    def create_dashboard_summary(
        self,
        all_anomalies: pd.DataFrame,
        municipality_scores: pd.DataFrame,
        summary_stats: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Create single-page dashboard visualization with 4 charts and key metrics.
        
        Creates a comprehensive dashboard with:
        - Top 10 municipalities (bar chart)
        - Anomaly distribution by type (pie chart)
        - Severity distribution (histogram)
        - Geographic heatmap (top regions)
        - Key metrics text boxes
        
        Args:
            all_anomalies: DataFrame with all detected anomalies
            municipality_scores: DataFrame with municipality scores
            summary_stats: Dictionary with summary statistics
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created dashboard image
        """
        self.logger.info("Creating dashboard summary visualization")
        
        # Generate filename with timestamp
        if filename is None:
            timestamp = self._get_timestamp()
            filename = f"dashboard_summary_{timestamp}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.png")
        
        # Get visualization config
        viz_config = self.config.get('visualization', {})
        dpi = viz_config.get('dpi', 300)
        
        # Create figure with 2x2 grid plus space for metrics
        fig = plt.figure(figsize=(20, 14))
        gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3, 
                             height_ratios=[0.15, 1, 1], 
                             width_ratios=[1, 1, 1])
        
        # Title and key metrics at the top (spanning all columns)
        ax_metrics = fig.add_subplot(gs[0, :])
        ax_metrics.axis('off')
        
        # Add title
        title_text = "EXECUTIVE DASHBOARD / ПАНЕЛЬ УПРАВЛЕНИЯ\nАнализ аномалий СберИндекс"
        ax_metrics.text(0.5, 0.85, title_text, 
                       ha='center', va='top', 
                       fontsize=20, fontweight='bold',
                       transform=ax_metrics.transAxes)
        
        # Add key metrics boxes
        total_anomalies = summary_stats.get('total_anomalies', 0)
        affected_municipalities = len(all_anomalies['territory_id'].unique()) if not all_anomalies.empty and 'territory_id' in all_anomalies.columns else 0
        critical_count = len(all_anomalies[all_anomalies['severity_score'] > 90]) if not all_anomalies.empty and 'severity_score' in all_anomalies.columns else 0
        avg_severity = all_anomalies['severity_score'].mean() if not all_anomalies.empty and 'severity_score' in all_anomalies.columns else 0
        
        # Create metrics boxes
        metrics = [
            ('Всего аномалий\nTotal Anomalies', f'{total_anomalies:,}', 'steelblue'),
            ('Критических\nCritical', f'{critical_count:,}', 'crimson'),
            ('Муниципалитетов\nMunicipalities', f'{affected_municipalities:,}', 'darkorange'),
            ('Средняя серьезность\nAvg Severity', f'{avg_severity:.1f}', 'forestgreen')
        ]
        
        box_width = 0.22
        box_start = 0.05
        box_spacing = 0.24
        
        for i, (label, value, color) in enumerate(metrics):
            x_pos = box_start + i * box_spacing
            
            # Draw box background
            from matplotlib.patches import Rectangle
            rect = Rectangle((x_pos, 0.1), box_width, 0.5, 
                           transform=ax_metrics.transAxes,
                           facecolor=color, alpha=0.2, 
                           edgecolor=color, linewidth=2)
            ax_metrics.add_patch(rect)
            
            # Add value (large)
            ax_metrics.text(x_pos + box_width/2, 0.45, value,
                          ha='center', va='center',
                          fontsize=24, fontweight='bold', color=color,
                          transform=ax_metrics.transAxes)
            
            # Add label (small)
            ax_metrics.text(x_pos + box_width/2, 0.2, label,
                          ha='center', va='center',
                          fontsize=10, color='black',
                          transform=ax_metrics.transAxes)
        
        # Chart 1: Top 10 Municipalities (top left)
        ax1 = fig.add_subplot(gs[1, 0:2])
        self._add_dashboard_top_municipalities(ax1, municipality_scores)
        
        # Chart 2: Anomaly Type Distribution (top right)
        ax2 = fig.add_subplot(gs[1, 2])
        self._add_dashboard_anomaly_types(ax2, summary_stats)
        
        # Chart 3: Severity Distribution (bottom left)
        ax3 = fig.add_subplot(gs[2, 0])
        self._add_dashboard_severity_distribution(ax3, all_anomalies)
        
        # Chart 4: Geographic Distribution (bottom middle and right)
        ax4 = fig.add_subplot(gs[2, 1:])
        self._add_dashboard_geographic_distribution(ax4, summary_stats)
        
        # Add timestamp footer
        timestamp_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        fig.text(0.99, 0.01, timestamp_text, ha='right', va='bottom', 
                fontsize=8, color='gray')
        
        # Save figure
        plt.savefig(filepath, dpi=dpi, bbox_inches='tight', facecolor='white')
        plt.close()
        
        self.logger.info(f"Saved dashboard summary to {filepath}")
        return filepath
    
    def _add_dashboard_top_municipalities(self, ax, municipality_scores: pd.DataFrame):
        """Add top 10 municipalities chart to dashboard."""
        if municipality_scores.empty:
            ax.text(0.5, 0.5, 'No Data Available', ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return
        
        # Get top 10 municipalities
        top_10 = municipality_scores.head(10).copy()
        
        # Prepare data (reverse for horizontal bar chart)
        municipalities = top_10['municipal_name'].tolist()[::-1]
        scores = top_10['total_severity_score'].tolist()[::-1]
        
        # Create horizontal bar chart
        bars = ax.barh(range(len(municipalities)), scores, color='coral', alpha=0.8, edgecolor='black')
        
        # Customize
        ax.set_xlabel('Total Severity Score', fontsize=11, fontweight='bold')
        ax.set_ylabel('Municipality', fontsize=11, fontweight='bold')
        ax.set_title('Top 10 Most Anomalous Municipalities', fontsize=13, fontweight='bold', pad=10)
        ax.set_yticks(range(len(municipalities)))
        ax.set_yticklabels(municipalities, fontsize=9)
        
        # Add value labels
        for i, (bar, score) in enumerate(zip(bars, scores)):
            width = bar.get_width()
            ax.text(width, bar.get_y() + bar.get_height()/2.,
                   f'{score:.0f}',
                   ha='left', va='center', fontsize=8, fontweight='bold')
        
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)
    
    def _add_dashboard_anomaly_types(self, ax, summary_stats: Dict[str, Any]):
        """Add anomaly type distribution pie chart to dashboard."""
        by_type = summary_stats.get('by_type', {})
        
        if not by_type:
            ax.text(0.5, 0.5, 'No Data Available', ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return
        
        # Prepare data
        labels = []
        sizes = []
        for anomaly_type, count in sorted(by_type.items(), key=lambda x: x[1], reverse=True):
            # Shorten labels for pie chart
            label_map = {
                'geographic_anomaly': 'Geographic',
                'cross_source_discrepancy': 'Cross-Source',
                'logical_inconsistency': 'Logical',
                'statistical_outlier': 'Statistical',
                'temporal_anomaly': 'Temporal',
                'data_quality_issue': 'Data Quality'
            }
            labels.append(label_map.get(anomaly_type, anomaly_type))
            sizes.append(count)
        
        # Create pie chart
        colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', '#ff99cc', '#c2c2f0']
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%',
                                           colors=colors[:len(sizes)],
                                           startangle=90, textprops={'fontsize': 9})
        
        # Make percentage text bold
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            autotext.set_fontsize(8)
        
        ax.set_title('Anomaly Distribution by Type', fontsize=13, fontweight='bold', pad=10)
    
    def _add_dashboard_severity_distribution(self, ax, all_anomalies: pd.DataFrame):
        """Add severity distribution histogram to dashboard."""
        if all_anomalies.empty or 'severity_score' not in all_anomalies.columns:
            ax.text(0.5, 0.5, 'No Data Available', ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return
        
        severity_scores = all_anomalies['severity_score'].dropna()
        
        if len(severity_scores) == 0:
            ax.text(0.5, 0.5, 'No Valid Data', ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return
        
        # Create histogram
        n, bins, patches = ax.hist(severity_scores, bins=15, 
                                   color='mediumseagreen', alpha=0.7,
                                   edgecolor='black', linewidth=1)
        
        # Color bars by severity level
        for i, patch in enumerate(patches):
            bin_center = (bins[i] + bins[i+1]) / 2
            if bin_center < 25:
                patch.set_facecolor('lightgreen')
            elif bin_center < 50:
                patch.set_facecolor('gold')
            elif bin_center < 75:
                patch.set_facecolor('orange')
            else:
                patch.set_facecolor('tomato')
        
        # Customize
        ax.set_xlabel('Severity Score', fontsize=11, fontweight='bold')
        ax.set_ylabel('Frequency', fontsize=11, fontweight='bold')
        ax.set_title('Severity Score Distribution', fontsize=13, fontweight='bold', pad=10)
        
        # Add threshold lines
        ax.axvline(x=25, color='green', linestyle='--', linewidth=1, alpha=0.5)
        ax.axvline(x=50, color='orange', linestyle='--', linewidth=1, alpha=0.5)
        ax.axvline(x=75, color='red', linestyle='--', linewidth=1, alpha=0.5)
        
        # Add statistics text
        stats_text = f'Mean: {severity_scores.mean():.1f}\nMedian: {severity_scores.median():.1f}'
        ax.text(0.98, 0.97, stats_text,
               transform=ax.transAxes,
               fontsize=9, verticalalignment='top', horizontalalignment='right',
               bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)
    
    def _add_dashboard_geographic_distribution(self, ax, summary_stats: Dict[str, Any]):
        """Add geographic distribution bar chart to dashboard."""
        by_region = summary_stats.get('by_region', {})
        
        if not by_region or len(by_region) == 0:
            ax.text(0.5, 0.5, 'No Regional Data Available', ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return
        
        # Sort and get top 15 regions
        sorted_regions = sorted(by_region.items(), key=lambda x: x[1], reverse=True)[:15]
        regions = [r[0] for r in sorted_regions]
        counts = [r[1] for r in sorted_regions]
        
        # Reverse for horizontal bar chart
        regions = regions[::-1]
        counts = counts[::-1]
        
        # Create horizontal bar chart
        bars = ax.barh(range(len(regions)), counts, color='steelblue', alpha=0.8, edgecolor='black')
        
        # Customize
        ax.set_xlabel('Number of Anomalies', fontsize=11, fontweight='bold')
        ax.set_ylabel('Region', fontsize=11, fontweight='bold')
        ax.set_title('Top 15 Regions by Anomaly Count', fontsize=13, fontweight='bold', pad=10)
        ax.set_yticks(range(len(regions)))
        ax.set_yticklabels(regions, fontsize=8)
        
        # Add value labels
        for i, (bar, count) in enumerate(zip(bars, counts)):
            width = bar.get_width()
            ax.text(width, bar.get_y() + bar.get_height()/2.,
                   f'{int(count)}',
                   ha='left', va='center', fontsize=8, fontweight='bold')
        
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)

    def generate_methodology_document(
        self,
        config: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Generate methodology document describing all detection methods used.
        
        Creates a comprehensive document explaining:
        - Overview of the anomaly detection system
        - Data sources and preprocessing steps
        - Detection methods for each anomaly type
        - Thresholds and parameters used
        - Interpretation guidelines
        
        Args:
            config: Configuration dictionary with thresholds and parameters
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created methodology document
        """
        self.logger.info("Generating methodology document")
        
        if filename is None:
            timestamp = self._get_timestamp()
            filename = f"methodology_{timestamp}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.md")
        
        # Build methodology content
        content = []
        content.append("# СберИндекс Anomaly Detection Methodology")
        content.append("")
        content.append(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append("")
        content.append("---")
        content.append("")

        # Overview section
        content.append("## 1. Overview")
        content.append("")
        content.append("This document describes the methodology used to detect anomalies in СберИндекс and Росстат data.")
        content.append("The system employs multiple detection methods to identify statistical outliers, temporal anomalies,")
        content.append("geographic deviations, cross-source discrepancies, and logical inconsistencies.")
        content.append("")
        content.append("### Objectives")
        content.append("")
        content.append("- Identify municipalities with unusual patterns in СберИндекс indicators")
        content.append("- Compare СберИндекс data with official Росстат statistics")
        content.append("- Detect data quality issues and logical inconsistencies")
        content.append("- Prioritize anomalies by severity for further investigation")
        content.append("")
        
        # Data sources section
        content.append("## 2. Data Sources")
        content.append("")
        content.append("### 2.1 СберИндекс Data")
        content.append("")
        content.append("- **Connection data** (connection.parquet): Municipal connectivity indicators")
        content.append("- **Consumption data** (consumption.parquet): Consumption patterns by category")
        content.append("- **Market access data** (market_access.parquet): Market accessibility indicators")
        content.append("")
        content.append("### 2.2 Росстат Data")
        content.append("")
        content.append("- **Population data** (2_bdmo_population.parquet): Population statistics by municipality")
        content.append("- **Migration data** (3_bdmo_migration.parquet): Migration flows and patterns")
        content.append("- **Salary data** (4_bdmo_salary.parquet): Average salary by industry and municipality")
        content.append("")
        content.append("### 2.3 Municipal Dictionary")
        content.append("")
        content.append("- **Municipal districts** (t_dict_municipal_districts.xlsx): Reference data for municipality names and identifiers")
        content.append("")

        # Detection methods section
        content.append("## 3. Detection Methods")
        content.append("")
        
        # Statistical outliers
        thresholds = config.get('thresholds', {})
        statistical = thresholds.get('statistical', {})
        
        content.append("### 3.1 Statistical Outlier Detection")
        content.append("")
        content.append("Identifies values that deviate significantly from statistical norms using three complementary methods:")
        content.append("")
        content.append(f"**Z-Score Method** (threshold: {statistical.get('z_score', 3.0)})")
        content.append("- Calculates how many standard deviations a value is from the mean")
        content.append(f"- Flags values with |z-score| > {statistical.get('z_score', 3.0)}")
        content.append("- Best for normally distributed data")
        content.append("")
        content.append(f"**IQR Method** (multiplier: {statistical.get('iqr_multiplier', 1.5)})")
        content.append("- Uses Interquartile Range (IQR = Q3 - Q1)")
        content.append(f"- Flags values < Q1 - {statistical.get('iqr_multiplier', 1.5)} × IQR or > Q3 + {statistical.get('iqr_multiplier', 1.5)} × IQR")
        content.append("- Robust to non-normal distributions")
        content.append("")
        content.append(f"**Percentile Method** (lower: {statistical.get('percentile_lower', 1)}%, upper: {statistical.get('percentile_upper', 99)}%)")
        content.append(f"- Flags values in bottom {statistical.get('percentile_lower', 1)}% or top {100 - statistical.get('percentile_upper', 99)}% of distribution")
        content.append("- Identifies extreme values regardless of distribution shape")
        content.append("")

        # Temporal anomalies
        temporal = thresholds.get('temporal', {})
        
        content.append("### 3.2 Temporal Anomaly Detection")
        content.append("")
        content.append("Detects unusual patterns over time:")
        content.append("")
        content.append(f"**Sudden Spikes** (spike threshold: {temporal.get('spike_threshold', 100)}%, drop threshold: {temporal.get('drop_threshold', -50)}%)")
        content.append("- Identifies period-over-period changes exceeding thresholds")
        content.append(f"- Growth rate > {temporal.get('spike_threshold', 100)}% indicates sudden spike")
        content.append(f"- Growth rate < {temporal.get('drop_threshold', -50)}% indicates sudden drop")
        content.append("")
        content.append("**Trend Reversals**")
        content.append("- Detects significant changes in trend direction")
        content.append("- Compares recent trend with historical trend")
        content.append("")
        content.append(f"**High Volatility** (multiplier: {temporal.get('volatility_multiplier', 2.0)})")
        content.append("- Calculates standard deviation of period-over-period changes")
        content.append(f"- Flags municipalities with volatility > {temporal.get('volatility_multiplier', 2.0)} × median volatility")
        content.append("")
        
        # Geographic anomalies
        geographic = thresholds.get('geographic', {})
        
        content.append("### 3.3 Geographic Anomaly Detection")
        content.append("")
        content.append("Identifies municipalities that differ from their geographic context:")
        content.append("")
        content.append(f"**Regional Outliers** (threshold: {geographic.get('regional_z_score', 2.0)} std dev)")
        content.append("- Calculates z-scores within each region/oblast")
        content.append(f"- Flags municipalities > {geographic.get('regional_z_score', 2.0)} standard deviations from regional mean")
        content.append("")
        content.append(f"**Cluster Outliers** (threshold: {geographic.get('cluster_threshold', 2.5)} std dev)")
        content.append("- Groups similar municipalities into clusters")
        content.append("- Identifies municipalities that differ significantly from their cluster")
        content.append("")
        content.append("**Urban vs Rural Analysis**")
        content.append("- Separate analysis for urban and rural municipalities")
        content.append("- Accounts for structural differences between municipality types")
        content.append("")

        # Cross-source comparison
        cross_source = thresholds.get('cross_source', {})
        
        content.append("### 3.4 Cross-Source Comparison")
        content.append("")
        content.append("Compares СберИндекс data with Росстат official statistics:")
        content.append("")
        content.append(f"**Correlation Analysis** (threshold: {cross_source.get('correlation_threshold', 0.5)})")
        content.append("- Calculates correlation coefficients between comparable indicators")
        content.append(f"- Flags indicator pairs with correlation < {cross_source.get('correlation_threshold', 0.5)}")
        content.append("- Low correlation may indicate data quality issues")
        content.append("")
        content.append(f"**Discrepancy Detection** (threshold: {cross_source.get('discrepancy_threshold', 50)}%)")
        content.append("- Calculates percentage difference between СберИндекс and Росстат values")
        content.append(f"- Flags municipalities with differences > {cross_source.get('discrepancy_threshold', 50)}%")
        content.append("- Ranks municipalities by magnitude of discrepancy")
        content.append("")
        
        # Logical consistency
        logical = thresholds.get('logical', {})
        
        content.append("### 3.5 Logical Consistency Checks")
        content.append("")
        content.append("Validates data for logical consistency:")
        content.append("")
        if logical.get('check_negative_values', True):
            content.append("**Negative Value Detection**")
            content.append("- Identifies negative values where only positive values are logically possible")
            content.append("- Examples: population, consumption amounts, salary")
            content.append("")
        
        if logical.get('check_impossible_ratios', True):
            content.append("**Impossible Ratio Detection**")
            content.append("- Detects logically impossible relationships")
            content.append("- Examples: consumption exceeding capacity, percentages > 100%")
            content.append("")
        
        content.append("**Contradictory Indicators**")
        content.append("- Identifies conflicting metrics")
        content.append("- Examples: high consumption with low connection rates")
        content.append("")

        # Severity scoring
        content.append("## 4. Severity Scoring")
        content.append("")
        content.append("Each anomaly is assigned a severity score from 0 to 100 based on:")
        content.append("")
        content.append("- **Deviation magnitude**: How far the value deviates from expected")
        content.append("- **Statistical significance**: Z-score or other statistical measures")
        content.append("- **Impact potential**: Importance of the affected indicator")
        content.append("")
        content.append("### Severity Categories")
        content.append("")
        content.append("| Score Range | Category | Interpretation |")
        content.append("|-------------|----------|----------------|")
        content.append("| 0-25 | Low | Minor deviation, may be normal variation |")
        content.append("| 25-50 | Medium | Notable deviation, worth investigating |")
        content.append("| 50-75 | High | Significant deviation, likely requires attention |")
        content.append("| 75-100 | Critical | Extreme deviation, requires immediate investigation |")
        content.append("")
        
        # Interpretation guidelines
        content.append("## 5. Interpretation Guidelines")
        content.append("")
        content.append("### 5.1 Understanding Anomalies")
        content.append("")
        content.append("An anomaly does not necessarily indicate an error. It may represent:")
        content.append("")
        content.append("- **Data quality issues**: Missing data, measurement errors, reporting problems")
        content.append("- **Genuine unusual patterns**: Real events or conditions that differ from the norm")
        content.append("- **Structural differences**: Unique characteristics of specific municipalities")
        content.append("- **Temporal events**: Seasonal effects, one-time events, policy changes")
        content.append("")

        content.append("### 5.2 Investigation Steps")
        content.append("")
        content.append("When investigating an anomaly:")
        content.append("")
        content.append("1. **Verify the data**: Check source data for accuracy and completeness")
        content.append("2. **Consider context**: Look for explanations in local conditions or events")
        content.append("3. **Compare indicators**: Check if multiple indicators show similar patterns")
        content.append("4. **Review temporal trends**: Examine historical data for patterns")
        content.append("5. **Geographic comparison**: Compare with neighboring municipalities")
        content.append("6. **Consult domain experts**: Seek input from local specialists")
        content.append("")
        
        content.append("### 5.3 False Positives")
        content.append("")
        content.append("Some anomalies may be false positives due to:")
        content.append("")
        content.append("- **Small sample sizes**: Limited data leading to statistical artifacts")
        content.append("- **Legitimate outliers**: Genuinely unique municipalities")
        content.append("- **Measurement differences**: Different methodologies between data sources")
        content.append("- **Timing differences**: Data collected at different times")
        content.append("")
        
        # Configuration summary
        content.append("## 6. Configuration Summary")
        content.append("")
        content.append("### Detection Thresholds")
        content.append("")
        content.append("```yaml")
        content.append("thresholds:")
        content.append("  statistical:")
        content.append(f"    z_score: {statistical.get('z_score', 3.0)}")
        content.append(f"    iqr_multiplier: {statistical.get('iqr_multiplier', 1.5)}")
        content.append(f"    percentile_lower: {statistical.get('percentile_lower', 1)}")
        content.append(f"    percentile_upper: {statistical.get('percentile_upper', 99)}")
        content.append("  temporal:")
        content.append(f"    spike_threshold: {temporal.get('spike_threshold', 100)}")
        content.append(f"    drop_threshold: {temporal.get('drop_threshold', -50)}")
        content.append(f"    volatility_multiplier: {temporal.get('volatility_multiplier', 2.0)}")
        content.append("  geographic:")
        content.append(f"    regional_z_score: {geographic.get('regional_z_score', 2.0)}")
        content.append(f"    cluster_threshold: {geographic.get('cluster_threshold', 2.5)}")
        content.append("  cross_source:")
        content.append(f"    correlation_threshold: {cross_source.get('correlation_threshold', 0.5)}")
        content.append(f"    discrepancy_threshold: {cross_source.get('discrepancy_threshold', 50)}")
        content.append("```")
        content.append("")
        
        content.append("---")
        content.append("")
        content.append("*This methodology document was automatically generated by the СберИндекс Anomaly Detection System.*")
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content))
        
        self.logger.info(f"Generated methodology document: {filepath}")
        return filepath

    def generate_example_cases(
        self,
        all_anomalies: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Generate document with example cases for each anomaly type.
        
        Selects representative examples from detected anomalies and provides
        potential explanations for each type of anomaly.
        
        Args:
            all_anomalies: DataFrame with all detected anomalies
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created examples document
        """
        self.logger.info("Generating example cases document")
        
        if filename is None:
            timestamp = self._get_timestamp()
            filename = f"example_cases_{timestamp}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.md")
        
        # Build content
        content = []
        content.append("# Anomaly Detection Example Cases")
        content.append("")
        content.append(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append("")
        content.append("This document provides example cases for each type of anomaly detected in the analysis,")
        content.append("along with potential explanations and interpretation guidance.")
        content.append("")
        content.append("---")
        content.append("")

        # Define anomaly types and their descriptions
        anomaly_types = {
            'statistical_outlier': {
                'title': 'Statistical Outliers',
                'description': 'Values that deviate significantly from statistical norms',
                'potential_causes': [
                    'Data entry errors or measurement issues',
                    'Genuinely unique municipality characteristics',
                    'Seasonal or temporary effects',
                    'Recent policy changes or interventions',
                    'Economic shocks or unusual events'
                ]
            },
            'temporal_anomaly': {
                'title': 'Temporal Anomalies',
                'description': 'Unusual patterns over time including spikes, drops, and volatility',
                'potential_causes': [
                    'One-time events (natural disasters, major projects)',
                    'Policy implementation or regulatory changes',
                    'Economic cycles or seasonal patterns',
                    'Data collection timing issues',
                    'Structural changes in local economy'
                ]
            },
            'geographic_anomaly': {
                'title': 'Geographic Anomalies',
                'description': 'Municipalities that differ significantly from their region or neighbors',
                'potential_causes': [
                    'Unique geographic or economic characteristics',
                    'Different industrial structure',
                    'Border effects or proximity to major cities',
                    'Local policy differences',
                    'Data quality variations by region'
                ]
            },
            'cross_source_discrepancy': {
                'title': 'Cross-Source Discrepancies',
                'description': 'Large differences between СберИндекс and Росстат data',
                'potential_causes': [
                    'Different measurement methodologies',
                    'Timing differences in data collection',
                    'Coverage differences (formal vs informal economy)',
                    'Data quality issues in one or both sources',
                    'Definitional differences for indicators'
                ]
            },
            'logical_inconsistency': {
                'title': 'Logical Inconsistencies',
                'description': 'Logically impossible or contradictory values',
                'potential_causes': [
                    'Data entry errors',
                    'Unit conversion mistakes',
                    'Misaligned data from different time periods',
                    'Incorrect data aggregation',
                    'System errors in data processing'
                ]
            },
            'data_quality_issue': {
                'title': 'Data Quality Issues',
                'description': 'Missing data, duplicates, or other quality problems',
                'potential_causes': [
                    'Incomplete data collection',
                    'Reporting gaps from municipalities',
                    'System errors or data pipeline issues',
                    'Duplicate records in source systems',
                    'Data format inconsistencies'
                ]
            }
        }

        # Generate examples for each anomaly type
        for anomaly_type, info in anomaly_types.items():
            content.append(f"## {info['title']}")
            content.append("")
            content.append(f"**Description:** {info['description']}")
            content.append("")
            
            # Get examples of this anomaly type
            if not all_anomalies.empty and 'anomaly_type' in all_anomalies.columns:
                type_anomalies = all_anomalies[all_anomalies['anomaly_type'] == anomaly_type]
                
                if not type_anomalies.empty:
                    # Select top 3 examples by severity
                    examples = type_anomalies.nlargest(3, 'severity_score') if 'severity_score' in type_anomalies.columns else type_anomalies.head(3)
                    
                    content.append("### Example Cases")
                    content.append("")
                    
                    for idx, (_, row) in enumerate(examples.iterrows(), 1):
                        content.append(f"#### Example {idx}")
                        content.append("")
                        content.append(f"- **Municipality:** {row.get('municipal_name', 'N/A')}")
                        content.append(f"- **Region:** {row.get('region_name', 'N/A')}")
                        content.append(f"- **Indicator:** {row.get('indicator', 'N/A')}")
                        content.append(f"- **Actual Value:** {row.get('actual_value', 'N/A')}")
                        
                        if pd.notna(row.get('expected_value')):
                            content.append(f"- **Expected Value:** {row.get('expected_value', 'N/A')}")
                        
                        if pd.notna(row.get('deviation')):
                            content.append(f"- **Deviation:** {row.get('deviation', 'N/A')}")
                        
                        if pd.notna(row.get('deviation_pct')):
                            content.append(f"- **Deviation %:** {row.get('deviation_pct', 'N/A'):.2f}%")
                        
                        content.append(f"- **Severity Score:** {row.get('severity_score', 'N/A')}")
                        content.append(f"- **Detection Method:** {row.get('detection_method', 'N/A')}")
                        content.append(f"- **Data Source:** {row.get('data_source', 'N/A')}")
                        
                        if pd.notna(row.get('description')):
                            content.append(f"- **Description:** {row.get('description', 'N/A')}")
                        
                        content.append("")
                else:
                    content.append("### Example Cases")
                    content.append("")
                    content.append("*No anomalies of this type were detected in the current analysis.*")
                    content.append("")
            else:
                content.append("### Example Cases")
                content.append("")
                content.append("*No anomaly data available.*")
                content.append("")

            # Add potential causes
            content.append("### Potential Causes and Explanations")
            content.append("")
            for cause in info['potential_causes']:
                content.append(f"- {cause}")
            content.append("")
            
            # Add investigation recommendations
            content.append("### Investigation Recommendations")
            content.append("")
            
            if anomaly_type == 'statistical_outlier':
                content.append("1. Verify the data source and collection method")
                content.append("2. Check for data entry errors or unit conversion issues")
                content.append("3. Research local conditions that might explain the outlier")
                content.append("4. Compare with historical data for the same municipality")
                content.append("5. Consult with local experts or administrators")
            elif anomaly_type == 'temporal_anomaly':
                content.append("1. Review historical timeline for significant events")
                content.append("2. Check for policy changes or interventions during the period")
                content.append("3. Verify data collection dates and methods")
                content.append("4. Look for seasonal patterns or cyclical effects")
                content.append("5. Compare with neighboring municipalities for similar patterns")
            elif anomaly_type == 'geographic_anomaly':
                content.append("1. Research unique characteristics of the municipality")
                content.append("2. Compare industrial and economic structure with region")
                content.append("3. Check for geographic factors (location, resources, infrastructure)")
                content.append("4. Review local policies and regulations")
                content.append("5. Verify data quality and completeness for the region")
            elif anomaly_type == 'cross_source_discrepancy':
                content.append("1. Compare methodologies used by СберИндекс and Росстат")
                content.append("2. Check timing of data collection for both sources")
                content.append("3. Verify indicator definitions are comparable")
                content.append("4. Look for coverage differences (e.g., formal vs informal)")
                content.append("5. Contact data providers for clarification if needed")
            elif anomaly_type == 'logical_inconsistency':
                content.append("1. Verify data entry and processing steps")
                content.append("2. Check for unit conversion errors")
                content.append("3. Ensure data from same time period")
                content.append("4. Review data aggregation methods")
                content.append("5. Flag for immediate correction if confirmed as error")
            elif anomaly_type == 'data_quality_issue':
                content.append("1. Identify the source of missing or duplicate data")
                content.append("2. Check data collection and reporting processes")
                content.append("3. Review data pipeline for system errors")
                content.append("4. Contact data providers about gaps")
                content.append("5. Document quality issues for future reference")
            
            content.append("")
            content.append("---")
            content.append("")
        
        # Add general notes
        content.append("## General Notes")
        content.append("")
        content.append("### Interpreting Examples")
        content.append("")
        content.append("- Examples are selected based on severity scores and representativeness")
        content.append("- High severity does not always indicate errors - may be genuine unusual cases")
        content.append("- Multiple anomalies in the same municipality may be related")
        content.append("- Context and domain knowledge are essential for proper interpretation")
        content.append("")
        content.append("### Next Steps")
        content.append("")
        content.append("1. Review high-severity anomalies first")
        content.append("2. Group related anomalies by municipality or indicator")
        content.append("3. Prioritize based on impact and data quality concerns")
        content.append("4. Conduct detailed investigations for critical cases")
        content.append("5. Document findings and corrective actions")
        content.append("")
        content.append("---")
        content.append("")
        content.append("*This example cases document was automatically generated by the СберИндекс Anomaly Detection System.*")
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content))
        
        self.logger.info(f"Generated example cases document: {filepath}")
        return filepath

    def generate_readme(
        self,
        summary_stats: Dict[str, Any],
        output_files: Dict[str, str],
        filename: str = "README"
    ) -> str:
        """
        Generate README with interpretation instructions and usage guide.
        
        Creates a comprehensive README explaining:
        - Project overview
        - Output files and their contents
        - How to interpret results
        - Usage instructions
        
        Args:
            summary_stats: Dictionary with summary statistics
            output_files: Dictionary mapping file types to file paths
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the created README file
        """
        self.logger.info("Generating README document")
        
        filepath = os.path.join(self.output_dir, f"{filename}.md")
        
        # Build content
        content = []
        content.append("# СберИндекс Anomaly Detection Results")
        content.append("")
        content.append(f"**Analysis Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append("")
        content.append("---")
        content.append("")
        
        # Overview
        content.append("## Overview")
        content.append("")
        content.append("This directory contains the results of anomaly detection analysis performed on СберИндекс")
        content.append("and Росстат data. The analysis identifies unusual patterns, statistical outliers, and")
        content.append("potential data quality issues across Russian municipalities.")
        content.append("")
        
        # Key findings
        content.append("## Key Findings")
        content.append("")
        total_anomalies = summary_stats.get('total_anomalies', 0)
        total_municipalities = summary_stats.get('total_municipalities_affected', 0)
        
        content.append(f"- **Total Anomalies Detected:** {total_anomalies:,}")
        content.append(f"- **Municipalities Affected:** {total_municipalities:,}")
        
        severity_stats = summary_stats.get('severity_stats', {})
        if severity_stats:
            content.append(f"- **Average Severity Score:** {severity_stats.get('mean', 0):.2f}")
            content.append(f"- **Maximum Severity Score:** {severity_stats.get('max', 0):.2f}")
        
        content.append("")
        
        # Anomaly breakdown
        by_type = summary_stats.get('by_type', {})
        if by_type:
            content.append("### Anomalies by Type")
            content.append("")
            for anomaly_type, count in sorted(by_type.items(), key=lambda x: x[1], reverse=True):
                content.append(f"- **{anomaly_type}:** {count:,}")
            content.append("")

        # Output files
        content.append("## Output Files")
        content.append("")
        content.append("This analysis generated the following files:")
        content.append("")
        
        if output_files:
            for file_type, file_path in output_files.items():
                filename_only = os.path.basename(file_path)
                content.append(f"### {filename_only}")
                content.append("")
                
                if 'master' in file_type or 'csv' in filename_only.lower():
                    content.append("**Type:** CSV - Master anomalies table")
                    content.append("")
                    content.append("Contains all detected anomalies with detailed information:")
                    content.append("- Municipality identification and location")
                    content.append("- Indicator name and values")
                    content.append("- Anomaly type and detection method")
                    content.append("- Severity scores and deviations")
                    content.append("- Descriptions and timestamps")
                    content.append("")
                    content.append("**Use this file for:** Detailed analysis, filtering, and custom reporting")
                    
                elif 'summary' in file_type or 'xlsx' in filename_only.lower():
                    content.append("**Type:** Excel - Summary workbook with multiple sheets")
                    content.append("")
                    content.append("Contains organized views of anomalies:")
                    content.append("- **Overview:** Summary statistics and key findings")
                    content.append("- **Statistical_Outliers:** Statistical outlier anomalies")
                    content.append("- **Temporal_Anomalies:** Time-based anomalies")
                    content.append("- **Geographic_Anomalies:** Location-based anomalies")
                    content.append("- **Cross_Source_Discrepancies:** Data source comparison issues")
                    content.append("- **Logical_Inconsistencies:** Logic violation anomalies")
                    content.append("- **Top_Anomalous_Municipalities:** Ranked list of most anomalous municipalities")
                    content.append("- **Data_Dictionary:** Column definitions and explanations")
                    content.append("")
                    content.append("**Use this file for:** Quick overview, presentations, and stakeholder reports")
                    
                elif 'methodology' in file_type or 'methodology' in filename_only.lower():
                    content.append("**Type:** Markdown - Methodology documentation")
                    content.append("")
                    content.append("Describes the detection methods and analysis approach:")
                    content.append("- Data sources and preprocessing")
                    content.append("- Detection algorithms and thresholds")
                    content.append("- Severity scoring methodology")
                    content.append("- Interpretation guidelines")
                    content.append("")
                    content.append("**Use this file for:** Understanding how anomalies were detected")
                    
                elif 'example' in file_type or 'example' in filename_only.lower():
                    content.append("**Type:** Markdown - Example cases")
                    content.append("")
                    content.append("Provides representative examples for each anomaly type:")
                    content.append("- Real examples from the analysis")
                    content.append("- Potential causes and explanations")
                    content.append("- Investigation recommendations")
                    content.append("")
                    content.append("**Use this file for:** Learning to interpret anomalies")
                    
                elif '.png' in filename_only.lower():
                    content.append("**Type:** PNG - Visualization")
                    content.append("")
                    if 'distribution' in filename_only.lower():
                        content.append("Shows distribution of anomalies by type or severity")
                    elif 'municipalities' in filename_only.lower():
                        content.append("Shows top municipalities by anomaly severity")
                    elif 'heatmap' in filename_only.lower():
                        content.append("Shows geographic distribution of anomalies")
                    content.append("")
                    content.append("**Use this file for:** Presentations and visual analysis")
                
                content.append("")
        else:
            content.append("*File list will be populated after analysis completes.*")
            content.append("")

        # How to interpret results
        content.append("## How to Interpret Results")
        content.append("")
        
        content.append("### Understanding Severity Scores")
        content.append("")
        content.append("Each anomaly has a severity score from 0 to 100:")
        content.append("")
        content.append("| Score | Category | Action |")
        content.append("|-------|----------|--------|")
        content.append("| 0-25 | Low | Monitor, may be normal variation |")
        content.append("| 25-50 | Medium | Review when time permits |")
        content.append("| 50-75 | High | Investigate soon |")
        content.append("| 75-100 | Critical | Investigate immediately |")
        content.append("")
        
        content.append("### Understanding Anomaly Types")
        content.append("")
        content.append("**Statistical Outliers**")
        content.append("- Values that deviate significantly from statistical norms")
        content.append("- May indicate data errors or genuinely unusual cases")
        content.append("- Check data source and local context")
        content.append("")
        
        content.append("**Temporal Anomalies**")
        content.append("- Unusual changes over time (spikes, drops, volatility)")
        content.append("- May indicate events, policy changes, or data issues")
        content.append("- Review timeline and historical context")
        content.append("")
        
        content.append("**Geographic Anomalies**")
        content.append("- Municipalities that differ from their region")
        content.append("- May reflect unique characteristics or data quality issues")
        content.append("- Compare with neighboring municipalities")
        content.append("")
        
        content.append("**Cross-Source Discrepancies**")
        content.append("- Differences between СберИндекс and Росстат")
        content.append("- May indicate methodology differences or data quality issues")
        content.append("- Review both data sources and their methodologies")
        content.append("")
        
        content.append("**Logical Inconsistencies**")
        content.append("- Logically impossible or contradictory values")
        content.append("- Often indicate data errors requiring correction")
        content.append("- Prioritize for investigation and correction")
        content.append("")
        
        content.append("**Data Quality Issues**")
        content.append("- Missing data, duplicates, or format problems")
        content.append("- Indicate data collection or processing issues")
        content.append("- Work with data providers to resolve")
        content.append("")

        # Workflow recommendations
        content.append("## Recommended Workflow")
        content.append("")
        content.append("### 1. Start with the Overview")
        content.append("")
        content.append("- Open the Excel summary file")
        content.append("- Review the Overview sheet for key statistics")
        content.append("- Identify which anomaly types are most common")
        content.append("")
        
        content.append("### 2. Review Top Anomalous Municipalities")
        content.append("")
        content.append("- Check the Top_Anomalous_Municipalities sheet")
        content.append("- Focus on municipalities with highest severity scores")
        content.append("- Note patterns across multiple municipalities")
        content.append("")
        
        content.append("### 3. Investigate by Anomaly Type")
        content.append("")
        content.append("- Review each anomaly type sheet")
        content.append("- Start with Logical_Inconsistencies (likely errors)")
        content.append("- Then review Cross_Source_Discrepancies")
        content.append("- Finally examine Statistical_Outliers and others")
        content.append("")
        
        content.append("### 4. Use Visualizations")
        content.append("")
        content.append("- Review PNG visualizations for patterns")
        content.append("- Use geographic heatmap to identify regional issues")
        content.append("- Check severity distribution for overall data quality")
        content.append("")
        
        content.append("### 5. Deep Dive with CSV")
        content.append("")
        content.append("- Open master CSV for detailed analysis")
        content.append("- Filter and sort by specific criteria")
        content.append("- Export subsets for further investigation")
        content.append("")
        
        content.append("### 6. Consult Documentation")
        content.append("")
        content.append("- Read methodology document to understand detection methods")
        content.append("- Review example cases for interpretation guidance")
        content.append("- Use Data_Dictionary sheet for column definitions")
        content.append("")

        # Common questions
        content.append("## Common Questions")
        content.append("")
        
        content.append("### Q: Are all anomalies errors?")
        content.append("")
        content.append("**A:** No. Anomalies indicate unusual patterns that may be:")
        content.append("- Data quality issues or errors")
        content.append("- Genuine unusual cases (unique municipalities)")
        content.append("- Temporary events or seasonal effects")
        content.append("- Structural differences between municipalities")
        content.append("")
        content.append("Always investigate context before concluding an anomaly is an error.")
        content.append("")
        
        content.append("### Q: How should I prioritize investigations?")
        content.append("")
        content.append("**A:** Prioritize by:")
        content.append("1. Severity score (higher = more urgent)")
        content.append("2. Anomaly type (logical inconsistencies first)")
        content.append("3. Impact (important indicators and large municipalities)")
        content.append("4. Patterns (multiple related anomalies)")
        content.append("")
        
        content.append("### Q: What if I find many anomalies in one municipality?")
        content.append("")
        content.append("**A:** Multiple anomalies in one municipality may indicate:")
        content.append("- Systematic data quality issues for that municipality")
        content.append("- Genuinely unique characteristics")
        content.append("- Recent significant events or changes")
        content.append("")
        content.append("Investigate the municipality holistically rather than each anomaly separately.")
        content.append("")
        
        content.append("### Q: How do I report findings?")
        content.append("")
        content.append("**A:** For each investigated anomaly, document:")
        content.append("- Anomaly details (municipality, indicator, values)")
        content.append("- Investigation steps taken")
        content.append("- Root cause identified")
        content.append("- Corrective action needed (if any)")
        content.append("- Status (confirmed error, false positive, genuine outlier)")
        content.append("")
        
        content.append("### Q: Can I re-run the analysis?")
        content.append("")
        content.append("**A:** Yes. The analysis is reproducible:")
        content.append("1. Ensure all data files are in place")
        content.append("2. Run `python main.py` from the project root")
        content.append("3. New results will be generated with timestamps")
        content.append("4. Configuration can be adjusted in `config.yaml`")
        content.append("")

        # Technical details
        content.append("## Technical Details")
        content.append("")
        
        content.append("### Data Sources")
        content.append("")
        content.append("- **СберИндекс:** connection.parquet, consumption.parquet, market_access.parquet")
        content.append("- **Росстат:** 2_bdmo_population.parquet, 3_bdmo_migration.parquet, 4_bdmo_salary.parquet")
        content.append("- **Municipal Dictionary:** t_dict_municipal_districts.xlsx")
        content.append("")
        
        content.append("### Detection Methods")
        content.append("")
        content.append("The analysis uses multiple statistical and logical methods:")
        content.append("- Z-score analysis for statistical outliers")
        content.append("- IQR (Interquartile Range) method")
        content.append("- Percentile-based detection")
        content.append("- Temporal pattern analysis")
        content.append("- Geographic comparison")
        content.append("- Cross-source correlation and discrepancy checks")
        content.append("- Logical consistency validation")
        content.append("")
        content.append("See the methodology document for detailed explanations.")
        content.append("")
        
        content.append("### Configuration")
        content.append("")
        content.append("Detection thresholds and parameters are configurable in `config.yaml`.")
        content.append("Key parameters include:")
        content.append("- Statistical thresholds (z-score, IQR multiplier, percentiles)")
        content.append("- Temporal thresholds (spike/drop percentages, volatility)")
        content.append("- Geographic thresholds (regional z-score)")
        content.append("- Cross-source thresholds (correlation, discrepancy percentage)")
        content.append("")
        
        content.append("## Support and Contact")
        content.append("")
        content.append("For questions about:")
        content.append("- **Methodology:** Review the methodology document")
        content.append("- **Interpretation:** Review the example cases document")
        content.append("- **Technical issues:** Check project documentation and logs")
        content.append("- **Data quality:** Contact data providers (СберИндекс, Росстат)")
        content.append("")
        
        content.append("---")
        content.append("")
        content.append("## Quick Reference")
        content.append("")
        content.append("### File Extensions")
        content.append("")
        content.append("- `.csv` - Comma-separated values (open in Excel, Python, R)")
        content.append("- `.xlsx` - Excel workbook (open in Excel, LibreOffice)")
        content.append("- `.md` - Markdown document (open in text editor, Markdown viewer)")
        content.append("- `.png` - Image file (open in image viewer, browser)")
        content.append("")
        
        content.append("### Key Columns in CSV/Excel")
        content.append("")
        content.append("- `territory_id` - Unique municipality identifier")
        content.append("- `municipal_name` - Municipality name")
        content.append("- `indicator` - Name of the metric showing anomaly")
        content.append("- `anomaly_type` - Type of anomaly detected")
        content.append("- `severity_score` - Severity from 0-100")
        content.append("- `actual_value` - Observed value")
        content.append("- `expected_value` - Expected value (if applicable)")
        content.append("- `deviation` - Absolute deviation")
        content.append("- `description` - Human-readable explanation")
        content.append("")
        
        content.append("---")
        content.append("")
        content.append(f"*Generated by СберИндекс Anomaly Detection System on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')}*")
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content))
        
        self.logger.info(f"Generated README document: {filepath}")
        return filepath


class ExecutiveSummaryGenerator:
    """
    Generates executive summary for management reporting.
    
    Provides methods to:
    - Calculate summary statistics
    - Identify top municipalities by risk
    - Generate key findings in Russian
    - Generate recommendations
    """
    
    def __init__(self, config: Dict[str, Any] = None):
        """
        Initialize the ExecutiveSummaryGenerator.
        
        Args:
            config: Configuration dictionary
        """
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self.config = config or {}
    
    def generate(self, anomalies_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Generate executive summary from anomalies data.
        
        Args:
            anomalies_df: DataFrame containing all detected anomalies
            
        Returns:
            Dictionary containing:
            - total_anomalies: Total number of anomalies
            - critical_count: Number of critical anomalies (severity > 90)
            - affected_municipalities: Number of unique municipalities affected
            - top_10_municipalities: List of top 10 municipalities by risk
            - key_findings: List of key findings in Russian
            - recommendations: List of recommendations in Russian
        """
        self.logger.info("Generating executive summary")
        
        if anomalies_df.empty:
            self.logger.warning("No anomalies to summarize")
            return {
                'total_anomalies': 0,
                'critical_count': 0,
                'affected_municipalities': 0,
                'top_10_municipalities': [],
                'key_findings': ['Аномалии не обнаружены'],
                'recommendations': ['Продолжить мониторинг данных']
            }
        
        # Calculate summary statistics
        total_anomalies = len(anomalies_df)
        critical_count = len(anomalies_df[anomalies_df['severity_score'] > 90]) if 'severity_score' in anomalies_df.columns else 0
        affected_municipalities = anomalies_df['territory_id'].nunique() if 'territory_id' in anomalies_df.columns else 0
        
        # Identify top 10 municipalities by risk
        top_10_municipalities = self._identify_top_municipalities(anomalies_df)
        
        # Generate key findings in Russian
        key_findings = self._generate_key_findings(anomalies_df, total_anomalies, critical_count, affected_municipalities)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(anomalies_df, critical_count, affected_municipalities)
        
        summary = {
            'total_anomalies': total_anomalies,
            'critical_count': critical_count,
            'affected_municipalities': affected_municipalities,
            'top_10_municipalities': top_10_municipalities,
            'key_findings': key_findings,
            'recommendations': recommendations
        }
        
        self.logger.info(f"Executive summary generated: {total_anomalies} anomalies, {critical_count} critical, {affected_municipalities} municipalities affected")
        
        return summary
    
    def _identify_top_municipalities(self, anomalies_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Identify top 10 municipalities by risk score.
        
        Risk score is calculated as the sum of severity scores for all anomalies
        in a municipality, with additional weight for critical anomalies.
        
        Args:
            anomalies_df: DataFrame containing all anomalies
            
        Returns:
            List of dictionaries with municipality information
        """
        self.logger.info("Identifying top 10 municipalities by risk")
        
        if 'territory_id' not in anomalies_df.columns or 'severity_score' not in anomalies_df.columns:
            self.logger.warning("Missing required columns for municipality ranking")
            return []
        
        # Group by municipality and calculate risk metrics
        municipality_groups = anomalies_df.groupby(['territory_id', 'municipal_name', 'region_name'])
        
        municipality_risks = []
        for (territory_id, municipal_name, region_name), group in municipality_groups:
            # Calculate risk score
            total_severity = group['severity_score'].sum()
            anomaly_count = len(group)
            critical_count = len(group[group['severity_score'] > 90])
            max_severity = group['severity_score'].max()
            avg_severity = group['severity_score'].mean()
            
            # Calculate weighted risk score (emphasize critical anomalies)
            risk_score = total_severity + (critical_count * 50)  # Bonus for critical anomalies
            
            # Get anomaly types
            anomaly_types = group['anomaly_type'].unique().tolist() if 'anomaly_type' in group.columns else []
            
            municipality_risks.append({
                'territory_id': territory_id,
                'municipal_name': municipal_name,
                'region_name': region_name,
                'risk_score': risk_score,
                'total_severity': total_severity,
                'anomaly_count': anomaly_count,
                'critical_count': critical_count,
                'max_severity': max_severity,
                'avg_severity': avg_severity,
                'anomaly_types': anomaly_types
            })
        
        # Sort by risk score and get top 10
        municipality_risks.sort(key=lambda x: x['risk_score'], reverse=True)
        top_10 = municipality_risks[:10]
        
        self.logger.info(f"Identified top 10 municipalities from {len(municipality_risks)} total")
        
        return top_10
    
    def _generate_key_findings(
        self,
        anomalies_df: pd.DataFrame,
        total_anomalies: int,
        critical_count: int,
        affected_municipalities: int
    ) -> List[str]:
        """
        Generate key findings in Russian based on anomaly data.
        
        Args:
            anomalies_df: DataFrame containing all anomalies
            total_anomalies: Total number of anomalies
            critical_count: Number of critical anomalies
            affected_municipalities: Number of affected municipalities
            
        Returns:
            List of key findings as strings in Russian
        """
        self.logger.info("Generating key findings")
        
        findings = []
        
        # Finding 1: Overall anomaly count
        findings.append(f"Обнаружено {total_anomalies:,} аномалий в {affected_municipalities:,} муниципальных образованиях")
        
        # Finding 2: Critical anomalies
        if critical_count > 0:
            critical_pct = (critical_count / total_anomalies * 100) if total_anomalies > 0 else 0
            findings.append(f"Критических аномалий (severity > 90): {critical_count:,} ({critical_pct:.1f}% от общего числа)")
        
        # Finding 3: Most common anomaly type
        if 'anomaly_type' in anomalies_df.columns:
            type_counts = anomalies_df['anomaly_type'].value_counts()
            if not type_counts.empty:
                most_common_type = type_counts.index[0]
                most_common_count = type_counts.iloc[0]
                most_common_pct = (most_common_count / total_anomalies * 100) if total_anomalies > 0 else 0
                
                # Translate anomaly type to Russian
                type_translations = {
                    'geographic_anomaly': 'географические аномалии',
                    'cross_source_discrepancy': 'расхождения между источниками',
                    'logical_inconsistency': 'логические несоответствия',
                    'statistical_outlier': 'статистические выбросы',
                    'temporal_anomaly': 'временные аномалии',
                    'data_quality_issue': 'проблемы качества данных'
                }
                type_name_ru = type_translations.get(most_common_type, most_common_type)
                findings.append(f"Наиболее распространенный тип: {type_name_ru} ({most_common_count:,} случаев, {most_common_pct:.1f}%)")
        
        # Finding 4: Most affected region
        if 'region_name' in anomalies_df.columns:
            region_counts = anomalies_df['region_name'].value_counts()
            if not region_counts.empty:
                most_affected_region = region_counts.index[0]
                most_affected_count = region_counts.iloc[0]
                findings.append(f"Наиболее затронутый регион: {most_affected_region} ({most_affected_count:,} аномалий)")
        
        # Finding 5: Most problematic indicator
        if 'indicator' in anomalies_df.columns:
            indicator_counts = anomalies_df['indicator'].value_counts()
            if not indicator_counts.empty:
                most_problematic_indicator = indicator_counts.index[0]
                most_problematic_count = indicator_counts.iloc[0]
                findings.append(f"Наиболее проблемный показатель: {most_problematic_indicator} ({most_problematic_count:,} аномалий)")
        
        # Finding 6: Data source analysis
        if 'data_source' in anomalies_df.columns:
            source_counts = anomalies_df['data_source'].value_counts()
            if len(source_counts) > 1:
                source_distribution = ', '.join([f"{src}: {cnt:,}" for src, cnt in source_counts.items()])
                findings.append(f"Распределение по источникам данных: {source_distribution}")
        
        # Finding 7: Severity distribution
        if 'severity_score' in anomalies_df.columns:
            high_severity_count = len(anomalies_df[anomalies_df['severity_score'] > 75])
            if high_severity_count > 0:
                high_severity_pct = (high_severity_count / total_anomalies * 100) if total_anomalies > 0 else 0
                findings.append(f"Аномалий высокой серьезности (severity > 75): {high_severity_count:,} ({high_severity_pct:.1f}%)")
        
        self.logger.info(f"Generated {len(findings)} key findings")
        
        return findings
    
    def _generate_recommendations(
        self,
        anomalies_df: pd.DataFrame,
        critical_count: int,
        affected_municipalities: int
    ) -> List[str]:
        """
        Generate recommendations in Russian based on anomaly patterns.
        
        Args:
            anomalies_df: DataFrame containing all anomalies
            critical_count: Number of critical anomalies
            affected_municipalities: Number of affected municipalities
            
        Returns:
            List of recommendations as strings in Russian
        """
        self.logger.info("Generating recommendations")
        
        recommendations = []
        
        # Recommendation 1: Critical anomalies
        if critical_count > 0:
            recommendations.append(f"Приоритетно проверить {critical_count:,} критических аномалий (severity > 90)")
        
        # Recommendation 2: Most common anomaly type
        if 'anomaly_type' in anomalies_df.columns:
            type_counts = anomalies_df['anomaly_type'].value_counts()
            if not type_counts.empty:
                most_common_type = type_counts.index[0]
                
                # Type-specific recommendations
                if most_common_type == 'geographic_anomaly':
                    recommendations.append("Рассмотреть возможность корректировки географических порогов для уменьшения ложных срабатываний")
                elif most_common_type == 'cross_source_discrepancy':
                    recommendations.append("Провести сверку данных между СберИндекс и Росстат для выявления систематических расхождений")
                elif most_common_type == 'logical_inconsistency':
                    recommendations.append("Проверить качество исходных данных и процедуры их сбора")
                elif most_common_type == 'data_quality_issue':
                    recommendations.append("Улучшить процессы контроля качества данных на этапе сбора")
        
        # Recommendation 3: High concentration in specific municipalities
        if 'territory_id' in anomalies_df.columns:
            municipality_counts = anomalies_df['territory_id'].value_counts()
            if not municipality_counts.empty:
                max_anomalies_per_municipality = municipality_counts.iloc[0]
                if max_anomalies_per_municipality > 10:
                    recommendations.append(f"Провести детальный аудит муниципалитетов с множественными аномалиями (более 10)")
        
        # Recommendation 4: Data source issues
        if 'data_source' in anomalies_df.columns:
            source_counts = anomalies_df['data_source'].value_counts()
            if len(source_counts) > 0:
                dominant_source = source_counts.index[0]
                dominant_count = source_counts.iloc[0]
                total = len(anomalies_df)
                if dominant_count / total > 0.7:  # More than 70% from one source
                    source_name = 'СберИндекс' if dominant_source == 'sberindex' else 'Росстат'
                    recommendations.append(f"Обратить внимание на качество данных из источника {source_name} ({dominant_count:,} аномалий)")
        
        # Recommendation 5: Widespread issues
        total_municipalities = anomalies_df['territory_id'].nunique() if 'territory_id' in anomalies_df.columns else 0
        if total_municipalities > 100:
            recommendations.append("Рассмотреть возможность автоматической настройки порогов (auto-tuning) для снижения количества ложных срабатываний")
        
        # Recommendation 6: General monitoring
        recommendations.append("Продолжить регулярный мониторинг и анализ аномалий для выявления новых паттернов")
        
        # Recommendation 7: Documentation
        recommendations.append("Документировать выявленные аномалии и принятые меры для формирования базы знаний")
        
        self.logger.info(f"Generated {len(recommendations)} recommendations")
        
        return recommendations
